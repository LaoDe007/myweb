from django.shortcuts import render, redirect

from web01.utils.pagination import Pagination
from web01 import models

from openpyxl import load_workbook


def fpl_list(request):
    queryset = models.FactoryPartsData.objects.all()
    # 分页
    page_object = Pagination(request, queryset)
    search_data = ""
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.page_html(),
        "search_data": search_data
    }

    return render(request, "fpl_list.html", context)


def fpl_multi(request):
    """批量上传（Excel文件）"""

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    if file_object:
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        # 3.循环获取每一行数据
        for row in sheet.iter_rows(min_row=2):
            part_universal = row[0].value
            identify_code = row[1].value
            part_code = row[2].value
            part_name = row[3].value
            plant_code = row[4].value
            supply_code = row[5].value
            supply_name = row[6].value
            person_purchase = row[7].value
            tpl_code = row[8].value
            tpl_name = row[9].value
            online_stock = row[10].value
            require_max = row[11].value
            require_one = row[12].value
            require_two = row[13].value
            require_three = row[14].value
            require_four = row[15].value
            require_five = row[16].value
            require_six = row[17].value
            require_seven = row[18].value
            exists = models.FactoryPartsData.objects.filter(identify_code=identify_code).exists()
            if not exists:
                models.FactoryPartsData.objects.create(part_universal=part_universal, identify_code=identify_code,
                                                       part_code=part_code, part_name=part_name, plant_code=plant_code,
                                                       supply_code=supply_code, supply_name=supply_name,
                                                       person_purchase=person_purchase, tpl_code=tpl_code,
                                                       tpl_name=tpl_name, require_max=require_max, online_stock=online_stock,
                                                       require_one=require_one, require_two=require_two,
                                                       require_three=require_three, require_four=require_four,
                                                       require_five=require_five, require_six=require_six,
                                                       require_seven=require_seven,
                                                       )

        return redirect("/fpl/list/")
    return redirect("/fpl/list/")


def tpl_delete(request, nid):
    """删除部门"""

    # 删除
    models.Warehouse.objects.filter(id=nid).delete()

    # 跳转回部门列表
    return redirect("/warehouse/list/")
