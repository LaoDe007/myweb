from django.shortcuts import render, redirect, HttpResponse
from django.core.files.uploadedfile import InMemoryUploadedFile
from web01 import models

from web01.utils.pagination import Pagination
from web01.utils.form import AdminModelForm, AdminEditModelForm,  AdminResetModelForm

from openpyxl import load_workbook


def admin_add(request):
    """添加管理员"""
    title = "新建管理员"
    if request.method == "GET":
        form = AdminModelForm()
        context = {
            "title": title,
            "form": form,
        }
        return render(request, "change.html", context)
    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        # 如果数据合法，保存到数据库
        form.save()
        return redirect("/admin/list/")
    context = {
        "title": title,
        "form": form,
    }
    # 校验失败，显示错误信息
    return render(request, 'change.html', context)


def admin_list(request):
    """管理员列表"""

    # 检查用户是否已经登录，已登录，继续向下运行；未登录，跳转回登录页面
    # 用户发来请求，获取cookie随机字符串，拿着随机字符串看看session中有没有
    # info = request.session.get('info')
    # if not info:
    #     return redirect("/login/")
    # info = request.session['info']
    #
    # # 搜索
    # data_dict = {}
    # search_data = request.GET.get("q", "")
    # if search_data:
    #     data_dict["username__contains"] = search_data
    #
    # queryset = models.Admin.objects.filter(**data_dict)
    queryset = models.Admin.objects.all()
    # 分页
    page_object = Pagination(request, queryset)
    search_data = ""
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.page_html(),
        "search_data": search_data
    }

    return render(request, "admin_list.html", context)


def admin_edit(request, nid):
    """编辑管理员"""
    # 对象/None
    row_object = models.Admin.objects.filter(id=nid).first()
    # 不存在
    if not row_object:
        return redirect("/admin/list/")
    title = "编辑管理员"

    if request.method == "GET":
        form = AdminEditModelForm(instance=row_object)
        context = {
            "title": title,
            "form": form,
        }
        return render(request, 'change.html', context)

    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    context = {
        "title": title,
        "form": form,
    }
    return render(request, 'change.html', context)


def admin_reset(request, nid):
    """重置密码"""
    # 对象/None
    row_object = models.Admin.objects.filter(id=nid).first()
    # 不存在
    if not row_object:
        return redirect("/admin/list/")
    title = "重置密码 - ".format(row_object.username)

    if request.method == "GET":
        form = AdminResetModelForm()
        context = {
            "title": title,
            "form": form,
        }
        return render(request, 'change.html', context)

    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    context = {
        "title": title,
        "form": form,
    }
    return render(request, 'change.html', context)


def admin_delete(request, nid):
    """删除部门"""

    # 删除
    models.Admin.objects.filter(id=nid).delete()

    # 跳转回部门列表
    return redirect("/admin/list/")


def admin_multi(request):
    """批量上传（Excel文件）"""

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Admin.objects.filter(username=text).exists()
        if not exists:
            models.Admin.objects.create(username=text)

    return redirect("/admin/list/")
