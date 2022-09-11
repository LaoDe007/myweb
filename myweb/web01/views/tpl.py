from django.shortcuts import render, redirect

from web01.utils.pagination import Pagination
from web01 import models

from openpyxl import load_workbook


def tpl_list(request):
    queryset = models.ThirdPartyLogisticsPartsData.objects.all()
    # 分页
    page_object = Pagination(request, queryset)
    search_data = ""
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.page_html(),
        "search_data": search_data
    }

    return render(request, "tpl_list.html", context)


def tpl_multi(request):
    """批量上传（Excel文件）"""

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    if file_object:
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        # 3.循环获取每一行数据
        for row in sheet.iter_rows(min_row=2):
            part_universal = row[0].value
            identify_code = row[1].value
            part_code = row[2].value
            part_name = row[3].value
            plant_code = row[4].value
            supply_code = row[5].value
            supply_name = row[6].value
            person_purchase = row[7].value
            tpl_code = row[8].value
            tpl_name = row[9].value
            parts_num = row[10].value
            exists = models.ThirdPartyLogisticsPartsData.objects.filter(identify_code=identify_code).exists()
            if not exists:
                models.ThirdPartyLogisticsPartsData.objects.create(part_universal=part_universal,
                                                                   identify_code=identify_code,
                                                                   part_code=part_code,
                                                                   part_name=part_name,
                                                                   plant_code=plant_code,
                                                                   supply_code=supply_code,
                                                                   supply_name=supply_name,
                                                                   person_purchase=person_purchase,
                                                                   tpl_code=tpl_code,
                                                                   tpl_name=tpl_name,
                                                                   parts_num=parts_num,
                                                                   )

        return redirect("/tpl/list/")
    return redirect("/tpl/list/")


def tpl_delete(request, nid):
    """删除部门"""

    # 删除
    models.Warehouse.objects.filter(id=nid).delete()

    # 跳转回部门列表
    return redirect("/warehouse/list/")


# def tpl_edit(request, nid):
#     """编辑仓库库存数据"""
#     row_object = models.Warehouse.objects.filter(id=nid).first()
#     if request.method == "GET":
#         # 根据nid，获取数据[obj,]
#         form = WarehouseEditModelForm(instance=row_object)
#         context = {
#             'form': form,
#         }
#         return render(request, "data_search.html", context)
#
#     # 获取用户提交的标题
#     parts_num = request.POST.get("parts_num")
#
#     # 根据ID找到数据库中的数据进行更新
#     models.Warehouse.objects.filter(id=nid).update(parts_num=parts_num)
#     # 跳转回部门列表
#     return redirect("/warehouse/list/")