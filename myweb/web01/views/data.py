from django.shortcuts import render, HttpResponse, redirect
from django import forms

from web01 import models
from web01.utils.pagination import Pagination
from web01.utils.form import InventoryDataModelForm
from openpyxl import load_workbook


def data_search(request):
    queryset = models.InventoryData.objects.all()
    # 分页
    page_object = Pagination(request, queryset)
    search_data = ""
    form = InventoryDataModelForm()
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.page_html(),
        "search_data": search_data,
        "form": form,
    }
    return render(request, "data_search.html", context)


def data_multi(request):
    """批量上传（Excel文件）"""

    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    if file_object:
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        # 3.循环获取每一行数据
        for row in sheet.iter_rows(min_row=2):

            store_warning = row[0].value
            online_warning = row[1].value
            days_store = float(row[2].value)
            part_universal = row[3].value
            identify_code = row[4].value
            part_code = row[5].value
            part_name = row[6].value
            plant_code = row[7].value
            supply_code = row[8].value
            supply_name = row[9].value
            person_purchase = row[10].value
            tpl_code = row[11].value
            tpl_name = row[12].value
            tpl_number = row[13].value
            online_stock = row[14].value
            require_max = row[15].value
            require_one = row[16].value
            require_two = row[17].value
            require_three = row[18].value
            require_four = row[19].value
            require_five = row[20].value
            require_six = row[21].value
            require_seven = row[22].value
            exists = models.InventoryData.objects.filter(identify_code=identify_code).exists()
            if not exists:
                models.InventoryData.objects.create(store_warning=store_warning, online_warning=online_warning,
                                                    days_store=days_store, tpl_number=tpl_number,
                                                    part_universal=part_universal, identify_code=identify_code,
                                                    part_code=part_code, part_name=part_name, plant_code=plant_code,
                                                    supply_code=supply_code, supply_name=supply_name,
                                                    person_purchase=person_purchase, tpl_code=tpl_code,
                                                    tpl_name=tpl_name, require_max=require_max,
                                                    online_stock=online_stock,
                                                    require_one=require_one, require_two=require_two,
                                                    require_three=require_three, require_four=require_four,
                                                    require_five=require_five, require_six=require_six,
                                                    require_seven=require_seven,
                                                    )

    return redirect("/data/search/")